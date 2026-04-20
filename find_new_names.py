#!/usr/bin/env python3
"""
find_new_names.py  (v4 — surname-anchored, 3-tier matching)
=============================================================
Extract BOLD multi-word names from a genealogy PDF and compare them
against a GEDCOM (.ged) database.

Matching rules:
  1. Single bold words are IGNORED (must be ≥2 bold words in a row)
  2. Surname tokens must overlap (accent-insensitive) — no surname = no match
  3. Given name tokens are compared after removing "common" tokens
     (auto-detected: tokens appearing in >30% of GED individuals, e.g. "Thi")
  4. Three-tier result:
       ✅ Strong  — surname matches + all qualifying given tokens match both ways
       ⚠️  Weak   — surname matches + some given tokens match, ≤1 extra each side
                   (user should review — could be middle-name variation or false positive)
       🔴 New     — no surname match, or no qualifying given token overlap

Usage:
    python find_new_names.py --ged family.ged --pdf genealogy.pdf --out results.xlsx

Options:
    --common-threshold FLOAT   Fraction of GED individuals a given-name token
                               must appear in to be excluded from matching
                               (default: 0.30 = 30%)

Requirements:
    pip install pdfplumber openpyxl python-gedcom
"""

import re
import sys
import unicodedata
import argparse
from pathlib import Path
from collections import defaultdict, Counter

try:
    import pdfplumber
except ImportError:
    sys.exit("Missing: pip install pdfplumber")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    sys.exit("Missing: pip install openpyxl")
try:
    from gedcom.parser import Parser
    from gedcom.element.individual import IndividualElement
except ImportError:
    sys.exit("Missing: pip install python-gedcom")


# ═══════════════════════════════════════════════════════════════════════════════
# 1. TEXT NORMALISATION
# ═══════════════════════════════════════════════════════════════════════════════

def strip_accents(s: str) -> str:
    """Remove diacritics: René→rene, Nguyễn→nguyen, Guéguen→gueguen."""
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    ).lower()

def clean_word(w: str) -> str:
    return w.strip(".,;:!?\"'»«()[]–—")

def token_set(name: str) -> set:
    """Accent-stripped token set, dropping single-char tokens."""
    return {strip_accents(clean_word(t)) for t in name.split()
            if len(clean_word(t)) > 1}


# ═══════════════════════════════════════════════════════════════════════════════
# 2. BOLD FONT DETECTION
# ═══════════════════════════════════════════════════════════════════════════════

def is_bold(fontname: str) -> bool:
    fn = (fontname or "").lower()
    return any(t in fn for t in ("bold", "-bd", ",bd", "heavy", "black",
                                  "demibold", "semibold"))

_PARTICLE_RE = re.compile(
    r"^(de|du|des|de la|d'|le|la|les|van|von|zu|af|of)$",
    re.IGNORECASE,
)


# ═══════════════════════════════════════════════════════════════════════════════
# 3. GEDCOM LOADING
# ═══════════════════════════════════════════════════════════════════════════════

def load_ged_names(ged_path: str) -> list:
    parser = Parser()
    parser.parse_file(ged_path, strict=False)
    people = []
    for element in parser.get_element_list():
        if not isinstance(element, IndividualElement):
            continue
        name_tuple = element.get_name()
        given   = (name_tuple[0] or "").strip()
        surname = (name_tuple[1] or "").strip()
        full    = f"{given} {surname}".strip()
        if not full:
            continue
        birth_data = element.get_birth_data()
        birth_year = ""
        if birth_data and birth_data[0]:
            m = re.search(r"\b(1[4-9]\d{2}|20\d{2})\b", birth_data[0])
            birth_year = m.group(1) if m else ""
        people.append({
            "full":             full,
            "given":            given,
            "surname":          surname,
            "birth_year":       birth_year,
            "_given_tokens":    token_set(given),
            "_surname_tokens":  token_set(surname),
        })
    return people


def build_common_given_tokens(ged_people: list, threshold: float) -> set:
    """
    Return given-name tokens appearing in more than `threshold` fraction
    of GED individuals. These are too common to be meaningful for matching
    (e.g. 'thi' in a Vietnamese database).
    """
    total = len(ged_people)
    if total == 0:
        return set()
    freq = Counter()
    for p in ged_people:
        for tok in p["_given_tokens"]:
            freq[tok] += 1
    return {tok for tok, cnt in freq.items() if cnt / total > threshold}


# ═══════════════════════════════════════════════════════════════════════════════
# 4. BOLD NAME EXTRACTION FROM PDF
# ═══════════════════════════════════════════════════════════════════════════════

def extract_bold_names(pdf_path: str) -> list:
    """
    Extract sequences of ≥2 consecutive bold words from every PDF page.
    Single bold words are silently discarded.
    Plain particles (de, du, le…) between two bold words are bridged.
    """
    candidates = defaultdict(lambda: {"count": 0, "pages": [], "raw": ""})

    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        print(f"  PDF has {total} pages — extracting bold multi-word names...")

        for page_num, page in enumerate(pdf.pages, 1):
            if page_num % 25 == 0:
                print(f"    ...page {page_num}/{total}")

            words = page.extract_words(
                extra_attrs=["fontname"],
                keep_blank_chars=False,
                x_tolerance=3,
                y_tolerance=3,
            )
            if not words:
                continue

            tagged = []
            for w in words:
                text = clean_word(w["text"])
                if not text:
                    continue
                tagged.append({
                    "text":     text,
                    "bold":     is_bold(w.get("fontname", "")),
                    "particle": bool(_PARTICLE_RE.match(text)),
                })

            i = 0
            while i < len(tagged):
                t = tagged[i]
                if t["bold"] and t["text"] and t["text"][0].isupper():
                    group = [t["text"]]
                    j = i + 1
                    while j < len(tagged):
                        nt = tagged[j]
                        if nt["bold"] and nt["text"] and nt["text"][0].isupper():
                            group.append(nt["text"])
                            j += 1
                        elif nt["particle"]:
                            if (j + 1 < len(tagged)
                                    and tagged[j + 1]["bold"]
                                    and tagged[j + 1]["text"][0].isupper()):
                                group.append(nt["text"])
                                j += 1
                            else:
                                break
                        else:
                            break

                    if len(group) >= 2:          # discard single-word groups
                        raw = " ".join(group)
                        key = strip_accents(raw)
                        candidates[key]["count"] += 1
                        if page_num not in candidates[key]["pages"]:
                            candidates[key]["pages"].append(page_num)
                        if not candidates[key]["raw"]:
                            candidates[key]["raw"] = raw
                    i = j
                else:
                    i += 1

    return [
        {"raw": d["raw"], "key": k, "count": d["count"], "pages": sorted(d["pages"])}
        for k, d in candidates.items()
    ]


# ═══════════════════════════════════════════════════════════════════════════════
# 5. MATCHING — 3-tier surname-anchored logic
# ═══════════════════════════════════════════════════════════════════════════════

STRONG = "✅ Strong match"
WEAK   = "⚠️  Weak match — review"
NEW    = "🔴 New — add to GED"

def _score_pair(pdf_given: set, ged_given: set) -> str:
    """
    Given two qualifying given-token sets (common tokens already removed),
    return STRONG, WEAK, or None (no match).
    """
    intersection = pdf_given & ged_given
    if not intersection:
        return None

    pdf_extra = pdf_given - ged_given   # tokens in PDF not in GED
    ged_extra = ged_given - pdf_given   # tokens in GED not in PDF

    if not pdf_extra and not ged_extra:
        return STRONG                   # perfect given-name match
    if len(pdf_extra) <= 1 and len(ged_extra) <= 1:
        return WEAK                     # at most 1 token differs on each side
    return None                         # too different


def find_best_match(pdf_name: str, ged_people: list, common_given: set) -> dict:
    """
    Find the best GED match for a PDF bold name.
    Priority: STRONG > WEAK > NEW.
    """
    pdf_tok   = token_set(pdf_name)
    pdf_given = pdf_tok - common_given  # we subtract surname per-candidate below

    best_status = None
    best_person = None

    for p in ged_people:
        # Step 1: surname must overlap
        if not (pdf_tok & p["_surname_tokens"]):
            continue

        # Step 2: compute qualifying given tokens (remove surname + common)
        pdf_given_q = pdf_tok - p["_surname_tokens"] - common_given
        ged_given_q = p["_given_tokens"] - common_given

        tier = _score_pair(pdf_given_q, ged_given_q)
        if tier is None:
            continue

        # Keep the best tier found
        if best_status is None or (tier == STRONG and best_status == WEAK):
            best_status = tier
            best_person = p

        if best_status == STRONG:
            break   # can't do better

    if best_person:
        return {"status": best_status, "best_match": best_person["full"]}
    return {"status": NEW, "best_match": ""}


# ═══════════════════════════════════════════════════════════════════════════════
# 6. EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════

STATUS_COLORS = {
    STRONG: "C6EFCE",   # green
    WEAK:   "FFEB9C",   # yellow
    NEW:    "FFC7CE",   # red
}

def write_excel(rows: list, out_path: str, ged_count: int,
                common_given: set, threshold: float):
    wb = openpyxl.Workbook()

    # ── Summary ───────────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum["A1"] = "Genealogy Name Comparison — v4 (surname-anchored, 3-tier)"
    ws_sum["A1"].font = Font(bold=True, size=13)
    ws_sum.merge_cells("A1:B1")
    ws_sum.column_dimensions["A"].width = 50
    ws_sum.column_dimensions["B"].width = 14

    stats      = Counter(r["status"] for r in rows)
    common_str = ", ".join(sorted(common_given)) or "(none detected)"

    for ri, (label, val) in enumerate([
        ("Names in GED database",          ged_count),
        ("Bold multi-word names in PDF",   len(rows)),
        ("", ""),
        (STRONG, stats[STRONG]),
        (WEAK,   stats[WEAK]),
        (NEW,    stats[NEW]),
        ("", ""),
        (f"Common given tokens excluded (>{threshold:.0%} of GED)", common_str),
    ], 3):
        ws_sum.cell(row=ri, column=1, value=label)
        ws_sum.cell(row=ri, column=2, value=val)
        if label in STATUS_COLORS:
            fill = PatternFill("solid", fgColor=STATUS_COLORS[label])
            ws_sum.cell(row=ri, column=1).fill = fill
            ws_sum.cell(row=ri, column=2).fill = fill
    ws_sum.cell(row=11, column=1).font = Font(italic=True, size=10)
    ws_sum.cell(row=11, column=2).font = Font(italic=True, size=10)

    # ── Legend ────────────────────────────────────────────────────────────────
    ws_leg = wb.create_sheet("Legend")
    ws_leg.column_dimensions["A"].width = 26
    ws_leg.column_dimensions["B"].width = 82
    ws_leg.cell(row=1, column=1, value="Status").font  = Font(bold=True)
    ws_leg.cell(row=1, column=2, value="Meaning").font = Font(bold=True)
    for ri, (status, meaning) in enumerate([
        (STRONG,
         "Surname matches AND all qualifying given-name tokens match both ways "
         "(accent-insensitive, order-insensitive). Almost certainly the same person."),
        (WEAK,
         "Surname matches and some given-name tokens overlap, but ≤1 token differs "
         "on each side. Could be a middle-name variation, a recording difference, "
         "or two different people sharing a common name. Review the GED match shown."),
        (NEW,
         "No surname match, or surname matched but no qualifying given-name token "
         "overlapped. Treat as a new individual — add to GED after verification."),
    ], 2):
        ws_leg.cell(row=ri, column=1, value=status).fill = \
            PatternFill("solid", fgColor=STATUS_COLORS[status])
        cell = ws_leg.cell(row=ri, column=2, value=meaning)
        cell.fill = PatternFill("solid", fgColor=STATUS_COLORS[status])
        cell.alignment = Alignment(wrap_text=True)
        ws_leg.row_dimensions[ri].height = 44

    # ── Results ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Results")
    headers = ["Status", "Bold name in PDF", "Occurrences",
               "Pages (first 10)", "Best GED match", "Your notes"]
    col_widths = [26, 30, 11, 32, 32, 32]
    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:F1"
    ws.row_dimensions[1].height = 30

    order = {NEW: 0, WEAK: 1, STRONG: 2}
    for ri, r in enumerate(
        sorted(rows, key=lambda r: (order[r["status"]], -r["count"])), 2
    ):
        pages_str = ", ".join(str(p) for p in r["pages"][:10])
        if len(r["pages"]) > 10:
            pages_str += f" … (+{len(r['pages'])-10} more)"
        fill = PatternFill("solid", fgColor=STATUS_COLORS[r["status"]])
        for c, v in enumerate(
            [r["status"], r["raw"], r["count"], pages_str, r["best_match"], ""], 1
        ):
            ws.cell(row=ri, column=c, value=v).fill = fill

    wb.save(out_path)
    print(f"\n✅  Results saved → {out_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# 7. MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="Find new genealogy names (bold, 2+ words) in a PDF vs a GEDCOM."
    )
    ap.add_argument("--ged", required=True, help="Path to your .ged file")
    ap.add_argument("--pdf", required=True, help="Path to your genealogy PDF")
    ap.add_argument("--out", default="new_names.xlsx",
                    help="Output Excel file (default: new_names.xlsx)")
    ap.add_argument("--common-threshold", type=float, default=0.30,
                    help="Given-name tokens appearing in more than this fraction "
                         "of GED individuals are excluded from matching "
                         "(default: 0.30 = 30%%)")
    args = ap.parse_args()

    for label, path in [("GED", args.ged), ("PDF", args.pdf)]:
        if not Path(path).exists():
            sys.exit(f"❌  {label} file not found: {path}")

    print(f"\n{'='*62}")
    print("  Genealogy Name Finder  v4")
    print("  Surname-anchored · common-token-aware · 3-tier output")
    print(f"{'='*62}")

    print(f"\n[1/4] Loading GEDCOM: {args.ged}")
    ged_people = load_ged_names(args.ged)
    print(f"      {len(ged_people)} individuals loaded")

    common_given = build_common_given_tokens(ged_people, args.common_threshold)
    if common_given:
        print(f"      Common given tokens excluded (>{args.common_threshold:.0%} of GED): "
              f"{', '.join(sorted(common_given))}")
    else:
        print(f"      No common given tokens above {args.common_threshold:.0%} threshold")

    print(f"\n[2/4] Extracting bold multi-word names from PDF: {args.pdf}")
    pdf_candidates = extract_bold_names(args.pdf)
    print(f"      {len(pdf_candidates)} unique bold name candidates found")

    print(f"\n[3/4] Matching against GED...")
    rows = [{**c, **find_best_match(c["raw"], ged_people, common_given)}
            for c in pdf_candidates]

    counts = Counter(r["status"] for r in rows)
    print(f"      {STRONG}: {counts.get(STRONG, 0)}")
    print(f"      {WEAK}:   {counts.get(WEAK, 0)}")
    print(f"      {NEW}:    {counts.get(NEW, 0)}")

    print(f"\n[4/4] Writing Excel: {args.out}")
    write_excel(rows, args.out, len(ged_people), common_given, args.common_threshold)

    print(f"\n{'='*62}")
    print(f"  🔴 {counts.get(NEW, 0)} new names to add to GED")
    print(f"  ⚠️  {counts.get(WEAK, 0)} weak matches to review manually")
    print(f"  ✅ {counts.get(STRONG, 0)} strong matches already in GED")
    print(f"{'='*62}\n")

if __name__ == "__main__":
    main()
