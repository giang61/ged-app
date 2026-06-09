#!/usr/bin/env python3
"""
process_results.py
==================
Post-processing pipeline for find_new_names.py output.

Steps:
  1. Load data/results.xlsx  (output of find_new_names.py)
  2. Split rows into three buckets: Strong matches, Weak matches, New adds
  3. Clean "Weak" and "New" names using the code-splitting logic from
     clean_names.py  (removes embedded genealogy ref codes like A.IV.3.b)
     One cell may expand into multiple rows after cleaning.
  4. Re-run the full surname-anchored matching against data/*.ged for every
     cleaned Weak and New name — some may now match after the noise is removed.
  5. Write three sheets to data/final_output.xlsx:
       • matches   — Strong matches (original) + newly promoted cleaned names
       • reviews   — Weak matches that are still weak after cleaning
       • new_adds  — Names with no match after cleaning

Usage:
    python process_results.py

All files are expected in the ./data/ folder.
Requires:  pip install pandas openpyxl python-gedcom

Matching logic (mirrors find_new_names.py v4):
  - Surname tokens must overlap (accent-insensitive)
  - Qualifying given-name tokens must overlap (common tokens auto-excluded)
  - Strong: all qualifying tokens match both ways (≤0 extras each side)
  - Weak:   surname matches + some overlap, ≤1 extra token each side
  - New:    no surname match, or no qualifying given-name overlap

Changes (v6):
  - Alias removal (prefix guard): "Quang Trung Nguyễn Huệ" tokens are
    stripped only when (a) all four tokens co-occur AND (b) other tokens
    remain after removal — i.e. the alias precedes another real name.
    A cell whose entire content is the alias is left untouched.
  - Reviews subset promotion: after re-matching, any review row whose token
    set (after alias removal) is a complete subset of a GED name's token set
    — or vice versa — is promoted directly to matches (Strong), regardless
    of word order.
"""

import re
import sys
import glob
import unicodedata
from pathlib import Path
from collections import defaultdict, Counter
from functools import lru_cache

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    sys.exit("Missing: pip install pandas openpyxl")

try:
    from gedcom.parser import Parser
    from gedcom.element.individual import IndividualElement
except ImportError:
    sys.exit("Missing: pip install python-gedcom")


DATA_DIR          = Path("./data")
RESULTS_FILE      = DATA_DIR / "results.xlsx"
OUTPUT_FILE       = DATA_DIR / "final_output.xlsx"
GED_FILE          = DATA_DIR / "nguyen.ged"    # path to your GEDCOM file
COMMON_THRESHOLD  = 0.30   # given-name tokens appearing in >30% of GED are excluded

# Status strings must match find_new_names.py v4
STRONG = "✅ Strong match"
WEAK   = "⚠️  Weak match — review"
NEW    = "🔴 New — add to GED"

STATUS_COLORS = {
    STRONG: "C6EFCE",
    WEAK:   "FFEB9C",
    NEW:    "FFC7CE",
}

# ═══════════════════════════════════════════════════════════════════════════════
# 1. NAME CLEANING  (from clean_names.py)
# ═══════════════════════════════════════════════════════════════════════════════

_CODE_RE = re.compile(
    r'[A-Za-z]\.(?:[IVXLCDM]+|\d+)(?:\.[0-9A-Za-z]+)*\.?'
)

def clean_text(text: str) -> str:
    text = str(text).replace("\u00A0", " ")
    text = _CODE_RE.sub(" ", text)
    text = re.sub(r'\.\s*', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def split_by_codes(cell: str) -> list:
    """Split a name cell on embedded ref codes, return list of clean name strings."""
    text = str(cell).replace("\u00A0", " ")
    chunks = _CODE_RE.split(text)
    results = []
    for chunk in chunks:
        name = clean_text(chunk)
        # Keep only chunks that look like they could be a name:
        # at least 2 words, at least one starts with a capital
        words = name.split()
        if len(words) >= 2 and any(w[0].isupper() for w in words if w):
            results.append(name)
    return results or [clean_text(cell)]   # fallback: return cleaned original


# ═══════════════════════════════════════════════════════════════════════════════
# 2. TEXT NORMALISATION  (mirrors find_new_names.py v4)
# ═══════════════════════════════════════════════════════════════════════════════

def strip_accents(s: str) -> str:
    """Return accent-stripped, lower-cased version of s."""
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    ).lower()   # .lower() ensures full case-insensitivity (feature #2)

def clean_word(w: str) -> str:
    return w.strip(".,;:!?\"'»«()[]–—")

def token_set(name: str) -> set:
    """
    Tokenise *name* into a set of normalised (accent-stripped, lower-cased)
    tokens, filtering out single-character noise.
    Case-insensitivity is guaranteed by strip_accents() calling .lower().
    """
    return {strip_accents(clean_word(t)) for t in name.split()
            if len(clean_word(t)) > 1}

def remove_alias_tokens(tokens: set) -> set:
    """
    Feature #1 — remove tokens belonging to the alias "Quang Trung Nguyễn Huệ"
    only when they appear as a *prefix* to another name, i.e. when other tokens
    remain after removal.

    Two guards are applied before any token is stripped:
      • All tokens of the alias phrase must be present in the set (so a lone
        "nguyen" in an unrelated name is never touched).
      • At least one non-alias token must remain after removal (so a cell
        whose *entire* content is the alias itself is left intact — it is a
        real name, not a prefix).
    """
    result = set(tokens)
    for alias_tokens in _ALIAS_TOKEN_SETS:
        if alias_tokens.issubset(result):
            remainder = result - alias_tokens
            if remainder:          # only strip when something else follows
                result = remainder
    return result

# ── Alias constants (defined here so strip_accents() is already available) ───
_ALIAS_PHRASES_RAW = [
    "Quang Trung Nguyễn Huệ",   # removed only when all four tokens co-occur
]
_ALIAS_TOKEN_SETS = [
    frozenset(strip_accents(w) for w in phrase.split())
    for phrase in _ALIAS_PHRASES_RAW
]
_ALIAS_PHRASES = [   # normalised strings for display in Summary sheet
    " ".join(strip_accents(w) for w in phrase.split())
    for phrase in _ALIAS_PHRASES_RAW
]


# ═══════════════════════════════════════════════════════════════════════════════
# 3. GEDCOM LOADING  (mirrors find_new_names.py v4)
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_name_tag(raw_value: str):
    """
    Robustly extract (given, surname) from a raw GEDCOM NAME tag value.

    python-gedcom's get_name() silently drops text that follows the closing
    slash when the surname comes first (e.g. '/Dương/ Công Kiên' → given='').
    This function always recovers all words regardless of slash placement:

      'Công Kiên /Dương/'   → given='Công Kiên',  surname='Dương'
      '/Dương/ Công Kiên'   → given='Công Kiên',  surname='Dương'
      'Công Kiên Dương'     → given='Kiên Dương',  surname='Công'  (first-word fallback)
    """
    raw = (raw_value or "").strip()
    m = re.search(r'/([^/]*)/', raw)
    if m:
        surname = m.group(1).strip()
        given   = re.sub(r'/[^/]*/', '', raw).strip()
    else:
        parts   = raw.split()
        surname = parts[0] if parts else ""
        given   = " ".join(parts[1:]) if len(parts) > 1 else ""
    return given, surname


def load_ged_names(ged_path: Path) -> list:
    parser = Parser()
    parser.parse_file(str(ged_path), strict=False)
    people = []
    for element in parser.get_element_list():
        if not isinstance(element, IndividualElement):
            continue
        # Use raw NAME tag value instead of get_name() which drops text
        # after the closing slash when the surname appears first.
        given, surname = "", ""
        for child in element.get_child_elements():
            if child.get_tag() == "NAME":
                given, surname = _parse_name_tag(child.get_value())
                break
        if not given and not surname:
            continue
        full = f"{given} {surname}".strip()
        people.append({
            "full":            full,
            "given":           given,
            "surname":         surname,
            # token sets are already lower-cased via strip_accents()
            "_given_tokens":   token_set(given),
            "_surname_tokens": token_set(surname),
            "_all_tokens":     token_set(full),   # used by order-independent fallback
        })
    return people

def build_common_given_tokens(ged_people: list, threshold: float) -> set:
    total = len(ged_people)
    if total == 0:
        return set()
    freq = Counter()
    for p in ged_people:
        for tok in p["_given_tokens"]:
            freq[tok] += 1
    return {tok for tok, cnt in freq.items() if cnt / total > threshold}


# ═══════════════════════════════════════════════════════════════════════════════
# 4. MATCHING  (mirrors find_new_names.py v4 + new features)
# ═══════════════════════════════════════════════════════════════════════════════

# ── Fuzzy token matching ──────────────────────────────────────────────────────

_FUZZY_MIN_LEN = 5   # only apply edit-distance to tokens >= 5 chars;
                     # shorter tokens (thi, van, le…) are too ambiguous

@lru_cache(maxsize=65536)
def _edit_distance(a: str, b: str) -> int:
    """Levenshtein edit distance, cached (same token pairs repeat across all names)."""
    if a == b:
        return 0
    if len(a) < len(b):
        a, b = b, a
    if len(a) - len(b) > 1:   # length gap > 1 → distance must be > 1, skip DP
        return 2
    prev = list(range(len(b) + 1))
    for ca in a:
        curr = [prev[0] + 1]
        for j, cb in enumerate(b, 1):
            curr.append(min(prev[j] + 1, curr[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = curr
    return prev[-1]

def _fuzzy_token_match(a: str, b: str) -> bool:
    """True if two tokens are equal or within edit distance 1 (long tokens only)."""
    if a == b:
        return True
    if len(a) >= _FUZZY_MIN_LEN and len(b) >= _FUZZY_MIN_LEN:
        return _edit_distance(a, b) <= 1
    return False

def _fuzzy_subtract(tokens: set, reference: set) -> set:
    """Return tokens that have NO fuzzy match in reference."""
    return {tok for tok in tokens
            if not any(_fuzzy_token_match(tok, ref) for ref in reference)}

def _fuzzy_intersect(set_a: set, set_b: set) -> bool:
    """True if at least one token in set_a fuzzy-matches any token in set_b."""
    return any(_fuzzy_token_match(a, b) for a in set_a for b in set_b)

def _fuzzy_matched_tokens(pdf_tok: set, reference: set) -> set:
    """Return the subset of pdf_tok that fuzzy-matches any token in reference."""
    return {pt for pt in pdf_tok
            if any(_fuzzy_token_match(pt, rt) for rt in reference)}


def _score_pair(pdf_given: set, ged_given: set) -> str | None:
    # If both sets are empty after common-token removal, all differentiating
    # tokens were common — the surname match alone is sufficient (Strong).
    if not pdf_given and not ged_given:
        return STRONG

    if not _fuzzy_intersect(pdf_given, ged_given):
        return None

    pdf_extra = _fuzzy_subtract(pdf_given, ged_given)
    ged_extra = _fuzzy_subtract(ged_given, pdf_given)
    if not pdf_extra and not ged_extra:
        return STRONG
    if len(pdf_extra) <= 1 and len(ged_extra) <= 1:
        return WEAK
    return None

def _score_flat(pdf_all: set, ged_all: set) -> str | None:
    """
    Feature #3 — order-independent fallback.
    Compare the *full* token sets (all words of the name, regardless of which
    are given vs surname) after removing alias tokens from both sides.
    Uses fuzzy matching so single-character typos do not block a match.
    """
    pdf_clean = remove_alias_tokens(pdf_all)
    ged_clean = remove_alias_tokens(ged_all)

    if not pdf_clean or not ged_clean:
        return None
    if not _fuzzy_intersect(pdf_clean, ged_clean):
        return None

    pdf_extra = _fuzzy_subtract(pdf_clean, ged_clean)
    ged_extra = _fuzzy_subtract(ged_clean, pdf_clean)
    if not pdf_extra and not ged_extra:
        return STRONG
    if len(pdf_extra) <= 1 and len(ged_extra) <= 1:
        return WEAK
    return None

def _is_subset_match(pdf_name: str, ged_people: list) -> dict | None:
    """
    Subset promotion check (feature — reviews → matches).

    After alias removal, if *either* token set is a subset of the other
    (in either direction), the names are considered a confirmed match and
    promoted to Strong.  The best (largest intersection) GED candidate is
    returned.  Uses fuzzy matching for subset containment.

    Returns {"status": STRONG, "best_match": <ged full name>} on success,
    or None if no subset relationship is found.
    """
    pdf_tok = remove_alias_tokens(token_set(pdf_name))
    if not pdf_tok:
        return None

    best_person = None
    best_overlap = 0

    for p in ged_people:
        ged_tok = remove_alias_tokens(p["_all_tokens"])
        if not ged_tok:
            continue
        # Fuzzy subset: every pdf token has a fuzzy match in ged, or vice versa
        pdf_in_ged = not _fuzzy_subtract(pdf_tok, ged_tok)
        ged_in_pdf = not _fuzzy_subtract(ged_tok, pdf_tok)
        if pdf_in_ged or ged_in_pdf:
            overlap = len(pdf_tok & ged_tok)
            if overlap > best_overlap:
                best_overlap = overlap
                best_person = p

    if best_person:
        return {"status": STRONG, "best_match": best_person["full"]}
    return None

def find_best_match(pdf_name: str, ged_people: list, common_given: set) -> dict:
    """
    Returns the best match dict {"status": ..., "best_match": ...}.

    Pass 1 — surname-anchored (with alias removal + fuzzy matching):
      Tokens from known aliases are stripped; surname overlap and given-name
      scoring both use fuzzy matching so single-character typos (e.g.
      Cexilia/Cecilia) in long tokens do not block a valid match.

    Pass 2 — flat / order-independent fallback:
      If Pass 1 found nothing, compare full token sets ignoring word order.
      Also fuzzy-aware.
    """
    pdf_tok       = token_set(pdf_name)
    pdf_tok_clean = remove_alias_tokens(pdf_tok)

    best_status = None
    best_person = None

    # ── Pass 1: surname-anchored ──────────────────────────────────────────────
    for p in ged_people:
        if not _fuzzy_intersect(pdf_tok_clean, p["_surname_tokens"]):
            continue
        matched_sur = _fuzzy_matched_tokens(pdf_tok_clean, p["_surname_tokens"])
        pdf_given_q = pdf_tok_clean - matched_sur - common_given
        ged_given_q = p["_given_tokens"] - common_given
        tier = _score_pair(pdf_given_q, ged_given_q)
        if tier is None:
            continue
        if best_status is None or (tier == STRONG and best_status == WEAK):
            best_status = tier
            best_person = p
        if best_status == STRONG:
            break

    if best_status == STRONG:
        return {"status": best_status, "best_match": best_person["full"]}

    # ── Pass 2: order-independent flat fallback ───────────────────────────────
    for p in ged_people:
        tier = _score_flat(pdf_tok, p["_all_tokens"])
        if tier is None:
            continue
        if best_status is None or (tier == STRONG and best_status != STRONG):
            best_status = tier
            best_person = p
        if best_status == STRONG:
            break

    if best_status == STRONG:
        return {"status": best_status, "best_match": best_person["full"]}

    # ── Pass 3: subset fallback ────────────────────────────────────────────────────────
    # Handles PDF “Nguyễn Bá Minh Triết” vs GED “An Phong Sô Bá Minh Triết Nguyễn”:
    # _score_flat fails because ged_extra has 3 tokens (an, phong, so).
    # But pdf_extra is empty — every PDF token IS in the GED name.
    #   pdf ⊆ ged  →  STRONG  (PDF is a short form of a longer GED name)
    #   ged ⊆ pdf  →  WEAK    (GED entry suspiciously short — flag for review)
    for p in ged_people:
        ged_all = p["_all_tokens"]
        pdf_not_in_ged = _fuzzy_subtract(pdf_tok, ged_all)
        ged_not_in_pdf = _fuzzy_subtract(ged_all, pdf_tok)
        if not pdf_not_in_ged:
            tier = STRONG
        elif not ged_not_in_pdf:
            tier = WEAK
        else:
            continue
        if best_status is None or (tier == STRONG and best_status != STRONG):
            best_status = tier
            best_person = p
        if best_status == STRONG:
            break

    if best_person:
        return {"status": best_status, "best_match": best_person["full"]}
    return {"status": NEW, "best_match": ""}


# ═══════════════════════════════════════════════════════════════════════════════
# 5. EXCEL WRITING
# ═══════════════════════════════════════════════════════════════════════════════

def _header_row(ws, headers: list, col_widths: list):
    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

def _write_sheet(ws, rows_df: pd.DataFrame, status_key: str, note_col: bool = True):
    """Write a DataFrame to a worksheet with colour coding."""
    fgcolor = STATUS_COLORS.get(status_key, "FFFFFF")

    cols = list(rows_df.columns)
    widths = []
    for col in cols:
        widths.append(max(14, min(40, max(len(str(col)), rows_df[col].astype(str).str.len().max() or 14) + 2)))
    if note_col:
        cols.append("Your notes")
        widths.append(32)

    _header_row(ws, cols, widths)
    fill = PatternFill("solid", fgColor=fgcolor)

    for ri, (_, row) in enumerate(rows_df.iterrows(), 2):
        for c, col in enumerate(cols, 1):
            val = row[col] if col in row.index else ""
            cell = ws.cell(row=ri, column=c, value=val)
            cell.fill = fill

    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(cols))}1"


def write_final_excel(matches_df, reviews_df, new_adds_df, output_path: Path,
                      ged_count: int, common_given: set):
    wb = openpyxl.Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum["A1"] = "Genealogy Name Processing — Final Output"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum.merge_cells("A1:B1")
    ws_sum.column_dimensions["A"].width = 46
    ws_sum.column_dimensions["B"].width = 14

    common_str = ", ".join(sorted(common_given)) or "(none)"
    alias_str  = ", ".join(f'"{p}"' for p in _ALIAS_PHRASES)
    for ri, (label, val) in enumerate([
        ("Names in GED database",                                   ged_count),
        ("", ""),
        (f"✅  matches   (strong match, no action needed)",          len(matches_df)),
        (f"⚠️   reviews   (weak match, verify manually)",            len(reviews_df)),
        (f"🔴  new_adds  (no match, add to GED)",                    len(new_adds_df)),
        ("", ""),
        (f"Common given tokens excluded (>{COMMON_THRESHOLD:.0%} of GED)", common_str),
        ("Alias phrases removed before matching",                    alias_str),
    ], 3):
        ws_sum.cell(row=ri, column=1, value=label)
        ws_sum.cell(row=ri, column=2, value=val)
        for status, color in [
            ("matches",  STATUS_COLORS[STRONG]),
            ("reviews",  STATUS_COLORS[WEAK]),
            ("new_adds", STATUS_COLORS[NEW]),
        ]:
            if status in str(label):
                fill = PatternFill("solid", fgColor=color)
                ws_sum.cell(row=ri, column=1).fill = fill
                ws_sum.cell(row=ri, column=2).fill = fill

    # ── Three data sheets ─────────────────────────────────────────────────────
    sheet_specs = [
        ("matches",  matches_df,  STRONG),
        ("reviews",  reviews_df,  WEAK),
        ("new_adds", new_adds_df, NEW),
    ]
    for sheet_name, df, status_key in sheet_specs:
        ws = wb.create_sheet(sheet_name)
        if df.empty:
            ws["A1"] = f"No entries in this category."
        else:
            _write_sheet(ws, df, status_key)

    wb.save(output_path)
    print(f"\n✅  Final output saved → {output_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# 6. MAIN PIPELINE
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    # ── Validate inputs ───────────────────────────────────────────────────────
    if not RESULTS_FILE.exists():
        sys.exit(f"❌  Not found: {RESULTS_FILE}")

    ged_path = GED_FILE
    if not ged_path.exists():
        sys.exit(f"❌  Not found: {ged_path}")

    print(f"\n{'='*62}")
    print("  Genealogy Results Processor  (v6)")
    print(f"{'='*62}")
    print(f"  Features active:")
    print(f"    • Alias removal  : {', '.join(repr(p) for p in _ALIAS_PHRASES_RAW)}")
    print(f"    •                  (removed only when all tokens co-occur)")
    print(f"    • Case-insensitive comparison enforced throughout")
    print(f"    • Order-independent flat-token fallback matching")
    print(f"    • Reviews subset promotion (either direction) → matches")

    # ── Step 1: Load results.xlsx ─────────────────────────────────────────────
    print(f"\n[1/5] Loading {RESULTS_FILE.name}...")
    df = pd.read_excel(RESULTS_FILE, sheet_name="Results")

    status_col = df.columns[0]
    name_col   = df.columns[1]

    # Normalise status values (strip whitespace/emoji variations)
    df[status_col] = df[status_col].astype(str).str.strip()

    strong_mask = df[status_col].str.contains("Strong", na=False)
    weak_mask   = df[status_col].str.contains("Weak",   na=False)
    new_mask    = df[status_col].str.contains("New",    na=False)

    strong_df = df[strong_mask].copy()
    weak_df   = df[weak_mask].copy()
    new_df    = df[new_mask].copy()

    print(f"      Loaded {len(df)} rows: "
          f"{len(strong_df)} strong · {len(weak_df)} weak · {len(new_df)} new")

    # ── Step 2: Load GED ──────────────────────────────────────────────────────
    print(f"\n[2/5] Loading GED: {ged_path.name}...")
    ged_people   = load_ged_names(ged_path)
    common_given = build_common_given_tokens(ged_people, COMMON_THRESHOLD)
    print(f"      {len(ged_people)} individuals loaded")
    if common_given:
        print(f"      Common given tokens excluded (>{COMMON_THRESHOLD:.0%}): "
              f"{', '.join(sorted(common_given))}")

    # ── Step 3: Clean & expand Weak + New rows ────────────────────────────────
    print(f"\n[3/5] Cleaning and expanding Weak + New names...")

    def expand_and_clean(source_df: pd.DataFrame) -> pd.DataFrame:
        """Apply split_by_codes to name column; expand one row → multiple rows."""
        expanded = []
        for _, row in source_df.iterrows():
            cleaned_names = split_by_codes(str(row[name_col]))
            for name in cleaned_names:
                new_row = row.copy()
                new_row[name_col] = name
                expanded.append(new_row)
        return pd.DataFrame(expanded).reset_index(drop=True) if expanded \
               else pd.DataFrame(columns=source_df.columns)

    weak_clean = expand_and_clean(weak_df)
    new_clean  = expand_and_clean(new_df)

    print(f"      {len(weak_df)} weak → {len(weak_clean)} rows after cleaning "
          f"(+{len(weak_clean)-len(weak_df)} from splits)")
    print(f"      {len(new_df)} new  → {len(new_clean)} rows after cleaning "
          f"(+{len(new_clean)-len(new_df)} from splits)")

    # ── Step 4: Re-match cleaned rows ─────────────────────────────────────────
    print(f"\n[4/5] Re-matching cleaned names against GED...")

    def rematch(source_df: pd.DataFrame) -> pd.DataFrame:
        if source_df.empty:
            return source_df
        results = source_df.copy()
        new_statuses   = []
        new_best_match = []
        for _, row in source_df.iterrows():
            match = find_best_match(str(row[name_col]), ged_people, common_given)
            new_statuses.append(match["status"])
            new_best_match.append(match["best_match"])
        results[status_col] = new_statuses
        if "Best GED match" in results.columns:
            results["Best GED match"] = new_best_match
        return results

    weak_rematched = rematch(weak_clean)
    new_rematched  = rematch(new_clean)

    # ── Step 5: Redistribute into final buckets ───────────────────────────────
    print(f"\n[5/5] Redistributing into final buckets...")

    def split_by_status(df: pd.DataFrame):
        s = df[status_col].str.contains("Strong", na=False)
        w = df[status_col].str.contains("Weak",   na=False)
        n = df[status_col].str.contains("New",    na=False)
        return df[s].copy(), df[w].copy(), df[n].copy()

    w_strong, w_weak, w_new = split_by_status(weak_rematched)
    n_strong, n_weak, n_new = split_by_status(new_rematched)

    # Combine weak survivors into reviews before subset promotion
    reviews_candidate = pd.concat([w_weak, n_weak], ignore_index=True)

    # ── Subset promotion: reviews → matches ───────────────────────────────────
    # For every remaining review row, check whether either token set is a
    # subset of the other (after alias removal).  If so, treat it as a Strong
    # match regardless of the ≤1-extra-token threshold used earlier.
    if not reviews_candidate.empty:
        promoted_mask = []
        promoted_statuses   = []
        promoted_best_match = []
        for _, row in reviews_candidate.iterrows():
            result = _is_subset_match(str(row[name_col]), ged_people)
            if result:
                promoted_mask.append(True)
                promoted_statuses.append(result["status"])
                promoted_best_match.append(result["best_match"])
            else:
                promoted_mask.append(False)
                promoted_statuses.append(row[status_col])
                promoted_best_match.append(
                    row.get("Best GED match", "") if "Best GED match" in row.index else ""
                )

        reviews_candidate = reviews_candidate.copy()
        reviews_candidate[status_col] = promoted_statuses
        if "Best GED match" in reviews_candidate.columns:
            reviews_candidate["Best GED match"] = promoted_best_match

        subset_promoted = sum(promoted_mask)
        print(f"      {subset_promoted} review(s) promoted to Strong via subset match")
    else:
        subset_promoted = 0

    # Re-split after subset promotion
    rev_strong, rev_weak, _ = split_by_status(reviews_candidate)

    matches_df  = pd.concat([strong_df, w_strong, n_strong, rev_strong], ignore_index=True)
    reviews_df  = rev_weak.copy().reset_index(drop=True)
    new_adds_df = pd.concat([w_new, n_new], ignore_index=True)

    for df_ref in [matches_df, reviews_df, new_adds_df]:
        if not df_ref.empty and name_col in df_ref.columns:
            df_ref.sort_values(name_col, inplace=True)
            df_ref.reset_index(drop=True, inplace=True)

    promoted = len(w_strong) + len(n_strong) + subset_promoted
    print(f"      {promoted} names promoted to Strong after cleaning")
    print(f"\n      Final counts:")
    print(f"        ✅  matches:  {len(matches_df)}")
    print(f"        ⚠️   reviews:  {len(reviews_df)}")
    print(f"        🔴  new_adds: {len(new_adds_df)}")

    # ── Write output ──────────────────────────────────────────────────────────
    write_final_excel(matches_df, reviews_df, new_adds_df,
                      OUTPUT_FILE, len(ged_people), common_given)

    print(f"\n{'='*62}")
    print(f"  Output → {OUTPUT_FILE}")
    print(f"  Sheets: matches · reviews · new_adds")
    print(f"{'='*62}\n")


if __name__ == "__main__":
    main()
